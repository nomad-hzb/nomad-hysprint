__author__ = "Edgar Nandayapa"
__version__ = "v0.0.1 2023"

from glob import glob
import pandas as pd
import seaborn as sns
import operator
import os
import re
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
# from openpyxl.styles import Font
import warnings

matplotlib.use('module://ipympl.backend_nbagg')

warnings.filterwarnings("ignore", message=".*cannot be placed.*", category=UserWarning)


def find_and_list_files(folder_path):
    file_patterns = ["JV_*.txt", "**/*JV_*.csv", "**/*JV_*.txt"]
    file_list = []
    for pattern in file_patterns:
        file_list.extend(glob(os.path.join(folder_path, pattern), recursive=True))
    file_list = list(set(file_list))

    return file_list


def load_files(file_list):
    # Consolidate file patterns for searching
    file_list.sort(key=natural_keys)
    # file_list = sorted(file_list, key=lambda x: int(x.split('\\')[-1].replace('JV_', '').replace('.txt', '')))

    # Initialize empty DataFrames for merged results
    jv_chars_merged = pd.DataFrame()
    curves_merged = pd.DataFrame()

    # Process each file
    for file_path in file_list:
        try:
            # Extract JV Characteristics and JV curve from the file
            jv_chars, jv_curve = process_file(file_path)
            # Merge data into cumulative DataFrames
            jv_chars_merged = pd.concat([jv_chars_merged, jv_chars],
                                        ignore_index=True) if not jv_chars.empty else jv_chars_merged
            curves_merged = pd.concat([curves_merged, jv_curve]) if not jv_curve.empty else curves_merged

        except Exception as e:  # Catch all exceptions to avoid stopping the loop
            print(f"Error processing {file_path}: {e}")
    curves_merged = curves_merged.reset_index()
    # Check if data was successfully loaded
    if jv_chars_merged.empty and curves_merged.empty:
        print("One of the files has an issue.")

    return jv_chars_merged, curves_merged


def replace_current_density_unit(idx):
    # This regular expression matches (mA/cm²) or (mA/cm^2) and captures the "mA/cm" part before the ² or ^2
    pattern = r'\(mA/cm(?:²|\^2)\)'
    replacement = '(mA/cm2)'
    return re.sub(pattern, replacement, idx)


def process_file(file_path):
    # Determines delimiter based on file extension
    linepos = find_separators_in_file(file_path)
    delimiter = '\t' if file_path.endswith('.txt') else ','

    try:
        # Initial attempt to read JV Characteristics
        jv_chars = pd.read_csv(file_path, skiprows=linepos[0], header=0, index_col=0, nrows=9,
                               delimiter=delimiter).transpose()
        # Attempt to read JV Curve - adjust parameters as per your file structure
        jv_curve = pd.read_csv(file_path, skiprows=linepos[1], header=0, index_col=None,
                               delimiter=delimiter).transpose()

        # Replace problematic character
        jv_chars.columns = [col.replace('²', '2') for col in jv_chars.columns]
        jv_curve.index = [replace_current_density_unit(idx) for idx in jv_curve.index]

        if not jv_chars.empty:
            jv_chars = add_extra_info(jv_chars, file_path, data_type='chars')

        if not jv_curve.empty:
            jv_curve = add_extra_info(jv_curve, file_path, data_type='curve')

    except pd.errors.EmptyDataError:
        jv_chars = pd.DataFrame()
        jv_curve = pd.DataFrame()

    return jv_chars, jv_curve


def add_extra_info(df, file_path, data_type):
    """
    Adds extra information to the DataFrame based on file path and data type.

    Parameters:
    - df: DataFrame to augment.
    - file_path: Path of the file being processed.
    - data_type: Type of data ('chars' for JV characteristics, 'curve' for JV curve).

    Returns:
    - DataFrame with added information.
    """
    norm_path = os.path.normpath(file_path)
    df['sample'] = file_path.split("JV_")[-1].rsplit(".", 1)[0]
    df['batch'] = norm_path.split(os.sep)[-2]
    df['condition'] = pd.NA

    split_index = df.index.to_series().str.split('_', expand=True)
    if data_type == "chars":
        df[['cell', 'direction', 'ilum']] = split_index

    if data_type == 'curve':
        # Assign the split results to new columns in df
        df[['variable', 'cell', 'direction', 'ilum']] = split_index

    return df


def find_separators_in_file(file_path):
    with open(file_path, "r") as file:
        lines = file.readlines()

    positions = []
    for index, line in enumerate(lines):
        if line.strip() == "--":
            positions.append(index + 1)
            # print(f"'--' found at line {index + 1}")

    return positions


def atoi(text):
    return int(text) if text.isdigit() else text


def natural_keys(text):
    return [atoi(c) for c in re.split(r'(\d+)', text)]


def name_by_condition(data, key_list, value_list):
    condition_dict = dict(zip(key_list, value_list))

    data["condition"] = data["sample"].map(condition_dict)

    return data


def data_filter_setup(df, filter_list):
    # Filter conditions
    # par = ["PCE(%)", "FF(%)", "FF(%)", "Voc(V)", "Jsc(mA/cm2)", "ilum"]
    # ope = ["<", "<", ">", "<", ">", "=="]
    # val = [40, 89, 24, 2, -30, "Light"]
    if not filter_list:
        filter_list = [("PCE(%)", "<", "40"), ("FF(%)", "<", "89"), ("FF(%)", ">", "24"), ("Voc(V)", "<", "2"),
                       ("Jsc(mA/cm2)", ">", "-30")]

    # List of operators
    operat = {"<": operator.lt, ">": operator.gt, "==": operator.eq,
              "<=": operator.le, ">=": operator.ge, "!=": operator.ne}

    data = df.copy()

    # Initialize the filter_reason column with empty strings
    data['filter_reason'] = ''
    filtering_options = []
    # for col, op, va in zip(par, ope, val):
    for col, op, va in filter_list:
        # Update the filter_reason for rows that do not meet the condition
        mask = operat[op](data[col], float(va))
        data.loc[~mask, 'filter_reason'] += f'{col} {op} {va}, '
        filtering_options.append(f'{col} {op} {va}')

    # Filter out rows that have any filter_reason
    trash = data[data['filter_reason'] != ''].copy()
    # Remove rows from data that were moved to trash
    data = data[data['filter_reason'] == '']
    # Clean up the filter_reason string by removing the trailing comma and space
    trash['filter_reason'] = trash['filter_reason'].str.rstrip(', ')

    print(f"\nThe {trash.shape[0]} samples listed here were filtered out based on the specified conditions: "
          f"{',  '.join(filtering_options)}.\n")
    print(trash[['sample', 'cell', 'filter_reason']].to_string(index=False))

    return data, trash, filtering_options


def jv_plot_curve_best(path, jvc, cur):
    # Find best device
    index_num = jvc["PCE(%)"].idxmax()
    sample = jvc.loc[index_num]["sample"]
    cell = jvc.loc[index_num]["cell"]

    # Filter data to focus on best device
    focus = cur.loc[(cur["sample"] == sample) & (cur["cell"] == cell)]  # &

    if len(focus) == 0:
        sample = jvc.loc[index_num]["identifier"]
        focus = cur.loc[(cur["sample"] == sample) & (cur["cell"] == cell)]  # &

    plotted = focus.copy().drop(
        columns=["index", "sample", "cell", "direction", "ilum", "batch", "condition"]).set_index(["variable"]).T
    dire = focus.loc[(focus["variable"] == "Voltage (V)")]["direction"].values
    ilum = focus.loc[(focus["variable"] == "Voltage (V)")]["ilum"].values

    fig, ax0 = plt.subplots(figsize=(10, 6))
    # Add axis now so that they are in the background
    ax0.axhline(linewidth=2, color='gray')
    ax0.axvline(linewidth=2, color='gray')

    # print(plotted)
    for c, p in enumerate(dire):
        x = plotted["Voltage (V)"].iloc[:, c]
        y = plotted["Current Density(mA/cm2)"].iloc[:, c]
        # print(p, x, y)

        # print(x,y)
        if dire[c] == "Reverse":
            ax0.plot(x, y, linestyle='-', marker='x', label=(dire[c] + "(" + ilum[c] + ")"))
        else:
            ax0.plot(x, y, linestyle='-', marker='.', label=(dire[c] + "(" + ilum[c] + ")"))
        ax0.legend(bbox_to_anchor=(1.02, 1), loc="upper left", prop={'size': 7.5})

    # Plot configuration
    ax0.set_xlabel('Voltage [V]', fontsize=13)
    ax0.set_ylabel('Current Density [mA/cm²]', fontsize=13)
    ax0.set_xlim([-0.2, 1.35])
    ax0.set_ylim([-25, 3])
    plt.grid(linestyle='dotted')

    # Get JV characteristics values and add them to variables
    df_rev = jvc.loc[(jvc["sample"] == sample) & (jvc["cell"] == cell)
                     & (jvc["direction"] == "Reverse")]
    df_for = jvc.loc[(jvc["sample"] == sample) & (jvc["cell"] == cell)
                     & (jvc["direction"] == "Forward")]

    char_vals = ['Voc(V)', 'Jsc(mA/cm2)', 'FF(%)', 'PCE(%)']

    char_rev = []
    char_for = []
    for cv in char_vals:
        char_rev.append(df_rev[cv].values[0])
        char_for.append(df_for[cv].values[0])

    # for mv in mpp_vals:
    v_f = df_for['V_mpp(V)'].values[0]
    v_r = df_rev['V_mpp(V)'].values[0]
    j_f = df_for['J_mpp(mA/cm2)'].values[0]
    j_r = df_rev['J_mpp(mA/cm2)'].values[0]

    ax0.plot(v_f, j_f, marker='.', color='r')
    ax0.plot(v_r, j_r, marker='X', color='r')

    text_bd_0 = f"\nVoc (V)\nJsc (mA/cm²)\nFF (%)\nPCE (%)"
    text_bd_1 = f"Rev\n{char_rev[0]:.2f}\n{char_rev[1]:.1f}\n{char_rev[2]:.1f}\n{char_rev[3]:.1f}"
    text_bd_2 = f"For\n{char_for[0]:.2f}\n{char_for[1]:.1f}\n{char_for[2]:.1f}\n{char_for[3]:.1f}"

    # Add jv information to the plot
    bottom_bdr = min(char_rev[1], char_for[1])
    right_bdr = max(char_rev[0], char_for[0])
    fsize = 15
    txt1 = ax0.text(right_bdr * 0.1, bottom_bdr * 0.2, text_bd_0, horizontalalignment='left',
                    verticalalignment="top", fontsize=fsize)
    tx1, ty1 = txt1.get_position()
    # print(tx1,ty1)
    ax0.text(tx1 + 0.60, ty1, text_bd_1, horizontalalignment='right', fontsize=fsize,
             verticalalignment="top", transform=plt.gca().transData)
    ax0.text(tx1 + 0.80, ty1, text_bd_2, horizontalalignment='right', fontsize=fsize,
             verticalalignment="top", transform=plt.gca().transData)

    # Add sample name
    ax0.text(0, 1, "Sample: " + sample + " (" + cell + ")", horizontalalignment='left',
             verticalalignment='bottom', transform=ax0.transAxes)

    #plt.tight_layout(rect=[0, 0, 1.25, 1])
    plt.subplots_adjust(right=0.8)

    sample_name = "JV_best_device.png"
    if not is_running_in_jupyter():
        plt.savefig(path + sample_name)
        plt.close()
        # plt.show()
        print(f"Saved JV curve of best device")
    return fig, sample_name


def jv_plot_by_cell_3x2(df, sample, path):
    # Filter the DataFrame for the specified sample
    focus = df[df["sample"] == sample]

    # Set up the matplotlib figure and axes for a 2x3 subplot layout
    fig, axs = plt.subplots(2, 3, figsize=(14, 10))
    axs = axs.flatten()  # Flatten the 2D array of axes for easy iteration

    # Group the DataFrame by 'cell'
    grouped = focus.groupby('cell')

    for i, (cell, group) in enumerate(grouped):
        if i >= len(axs):  # Check to prevent index error if there are more than 6 cells
            break

        # Extract 'Voltage (V)' and 'Current Density(mA/cm²)' for the cell
        # You might need to adjust this part based on your actual data structure
        for direction in group['direction'].unique():
            for ilum in group['ilum'].unique():
                plotted = group[(group["direction"] == direction) & (group["ilum"] == ilum)]
                plotted = plotted.copy().drop(
                    columns=["index", "sample", "cell", "direction", "ilum", "batch", "condition"]).set_index(
                    ["variable"]).T

                # Plot the curve
                label = f"{direction} ({ilum})"
                axs[i].plot(plotted["Voltage (V)"], plotted["Current Density(mA/cm2)"], label=label)

                # Set titles, labels, limits, and legends
                axs[i].set_title(f'Cell {cell}')
                axs[i].set_xlabel('Voltage [V]', fontsize=13)
                axs[i].set_ylabel('Current Density [mA/cm²]', fontsize=13)
                axs[i].set_xlim([-0.2, 1.35])
                axs[i].set_ylim([-25, 3])
                axs[i].axhline(linewidth=2, color='gray')
                axs[i].axvline(linewidth=2, color='gray')
                axs[i].grid(linestyle='dotted')
                axs[i].legend()

    fig.suptitle(f"Sample {sample}", fontsize=16)

    #plt.tight_layout()
    plt.subplots_adjust(top=0.92)

    image_name = f"JV_cells_by_sample_{sample}.png"
    if not is_running_in_jupyter():
        fig.savefig(path + image_name, dpi=300)
        plt.close()
        print(f"Saved JV_cells_by_sample_{sample}.png")

    return fig, image_name


def jv_plot_by_substrate(df, sample, path):
    focus = df[df["sample"] == sample]

    fig, ax = plt.subplots(figsize=(10, 6))

    grouped = focus.groupby('cell')

    color_cycle = plt.cm.rainbow(np.linspace(0, 1, grouped.ngroups))

    for i, ((cell, group), color) in enumerate(zip(grouped, color_cycle)):
        # Iterate through directions and illumination conditions
        for direction in group['direction'].unique():
            for ilum in group['ilum'].unique():
                plotted = group[(group["direction"] == direction) & (group["ilum"] == ilum)]
                plotted = plotted.copy().drop(
                    columns=["index", "sample", "cell", "direction", "ilum", "batch", "condition"]).set_index(
                    ["variable"]).T

                # Determine line style based on direction
                line_style = '-' if direction == 'Forward' else '--'

                # Plot the curve with unique color and appropriate line style
                label = f"Cell {cell}, {direction} ({ilum})"
                ax.plot(plotted["Voltage (V)"], plotted["Current Density(mA/cm2)"], label=label, color=color,
                        linestyle=line_style)

    # Set titles, labels, limits, and legends for the combined plot
    ax.set_xlabel('Voltage [V]', fontsize=13)
    ax.set_ylabel('Current Density [mA/cm²]', fontsize=13)
    ax.set_xlim([-0.2, 1.35])
    ax.set_ylim([-25, 3])
    ax.axhline(linewidth=2, color='gray')
    ax.axvline(linewidth=2, color='gray')
    ax.grid(linestyle='dotted')
    ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", prop={'size': 7.5})

    fig.suptitle(f"Sample {sample}", fontsize=16)

    #plt.tight_layout()
    #plt.subplots_adjust(top=0.92)
    plt.subplots_adjust(right=0.75)

    image_name = f"JV_combined_sample_{sample}.png"
    if not is_running_in_jupyter():
        fig.savefig(path + image_name, dpi=300)
        plt.close()
        print(f"Saved JV_combined_sample_{sample}.png")

    return fig, image_name


def jv_plot_together(df1, df2, path, namestring):
    # Prepare the data frame as before
    if not namestring == "All":
        df2_copy = pd.merge(df1[['cell', 'sample']], df2, on=['cell', 'sample'], how='inner').drop_duplicates()[df2.columns]
    else:
        df2_copy = df2.copy()

    df2_plot = drop_extra_cols_and_ready_to_plot(df2_copy)

    cols = []
    counters = {'Voltage (V)': 0, 'Current Density(mA/cm2)': 0}
    for col in df2_plot.columns:
        counters[col] += 1
        cols.append(f"{col} {counters[col]}")
    df2_plot.columns = cols

    # Start using the object-oriented approach here
    fig, ax = plt.subplots(figsize=(10, 6))

    num_pairs = len(df2_plot.columns) // 2
    for i in range(1, num_pairs + 1):
        if (df2_plot[f'Voltage (V) {i}'].diff()).all() > 0:
            line_style = 'solid'
        else:
            line_style = 'dashed'
        ax.plot(df2_plot[f'Voltage (V) {i}'], df2_plot[f'Current Density(mA/cm2) {i}'],
                color='dimgray', linestyle=line_style, linewidth=0.5)

    # Process the top devices as before
    top_df1 = df1.sort_values('PCE(%)', ascending=False).head(4)
    top_samples = top_df1.sort_values(by=['sample', 'batch', 'cell', 'direction'])[
        ['sample', 'batch', 'cell', 'direction']]
    best_devices = pd.merge(top_samples, df2, on=['sample', 'batch', 'cell', 'direction'], how='inner')

    best_plot = drop_extra_cols_and_ready_to_plot(best_devices)
    cols = []
    counters = {'Voltage (V)': 0, 'Current Density(mA/cm2)': 0}
    for col in best_plot.columns:
        counters[col] += 1
        cols.append(f"{col} {counters[col]}")
    best_plot.columns = cols

    selected_plot = best_devices[best_devices['variable'] == 'Voltage (V)']

    i = 1
    for _, row in selected_plot.iterrows():
        if row['direction'] == 'Forward':
            line_style = 'solid'
        else:
            line_style = 'dashed'
        ax.plot(best_plot[f'Voltage (V) {i}'], best_plot[f'Current Density(mA/cm2) {i}'],
                linestyle=line_style, linewidth=2,
                label=f"Sample {row['sample']}-cell {row['cell']}")
        i += 1

    # Add empty plots for legend entries
    ax.plot([], [], label='Forward', color='dimgray', linestyle="solid", linewidth=0.5)
    ax.plot([], [], label='Reverse', color='dimgray', linestyle="dashed", linewidth=0.5)

    # Set titles, labels, limits, and legend with the ax object
    ax.set_title(f'{namestring} curves')
    ax.set_xlabel('Voltage [V]', fontsize=13)
    ax.set_ylabel('Current Density [mA/cm²]', fontsize=13)
    ax.set_xlim([-0.2, 1.35])
    ax.set_ylim([-25, 3])
    ax.axhline(linewidth=2, color='gray')
    ax.axvline(linewidth=2, color='gray')
    ax.grid(linestyle='dotted')
    ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", prop={'size': 7.5})

    # Ensure the layout is tight
    #fig.tight_layout()
    plt.subplots_adjust(right=0.8)

    image_name = f"JV_together_{namestring}.png"
    if not is_running_in_jupyter():
        plt.savefig(path + image_name, dpi=300)
        plt.close()
        # plt.show()
        print(f"Saved JV_together_{namestring}.png")

    return fig, image_name


def drop_extra_cols_and_ready_to_plot(df):
    cols_to_remove = ["index", "sample", "cell", "direction", "ilum", "batch", "condition"]

    common_cols_to_remove = df.columns.intersection(cols_to_remove)
    df_clean = df.copy().drop(columns=common_cols_to_remove)

    df_clean = df_clean.set_index(["variable"]).T
    return df_clean


def boxplot_all_cells(path, wb, data, var_x, var_y, trash, datatype):
    names_dict = {
        "voc": 'Voc(V)', "jsc": 'Jsc(mA/cm2)', "ff": 'FF(%)', "pce": 'PCE(%)',
        "vmpp": 'V_mpp(V)', "jmpp": 'J_mpp(mA/cm2)', "pmpp": 'P_mpp(mW/cm2)',
        "rser": 'R_series(Ohmcm2)', "rshu": 'R_shunt(Ohmcm2)'
    }
    var_name_y = names_dict[var_y]

    try:
        data["sample"] = data["sample"].astype(int)
    except ValueError:
        pass

    data['Jsc(mA/cm2)'] = data['Jsc(mA/cm2)'].abs()

    # Plotting style and colors
    sns.set(style="whitegrid")

    # Calculate statistics and add median with respect to var_x
    descriptor = data.groupby(var_x)[var_name_y].describe()

    # Make sure trash[0] has data for the var_x categories
    trash_categories = trash[0][var_x].unique()
    trashcriptor = trash[0].groupby(var_x)[var_name_y].describe() if not trash[0].empty else pd.DataFrame()

    # Ordering
    order_parameter = "alphabetic"  # Can be 'count', 'mean', 'std', 'min', 'max', 'median', 'alphabetic'
    if order_parameter != "alphabetic":
        orderc = descriptor.sort_values(by=[order_parameter])["count"].index
    else:
        orderc = descriptor.sort_index()["count"].index

    # Create figure
    fig, ax = plt.subplots(figsize=(10, 6))

    # Boxplot
    sns.boxplot(x=var_x, y=var_name_y, data=data, showfliers=False, showmeans=True,
                order=orderc, meanprops={"marker": "*", "markerfacecolor": "white",
                                         "markeredgecolor": "black", "markersize": "10"}, ax=ax)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=15)  # Adjust rotation if necessary

    # Swarmplot
    sns.swarmplot(x=var_x, y=var_name_y, data=data, color=".1", alpha=0.4, order=orderc, ax=ax)

    # Create a dictionary to map categories to their counts
    data_counts = data.groupby(var_x)[var_name_y].count().to_dict()
    trash_counts = trash[0].groupby(var_x)[var_name_y].count().to_dict() if not trash[0].empty else {}

    # Calculate y-axis limits for text placement
    y_min = data[var_name_y].min()
    y_max = data[var_name_y].max()
    y_range = y_max - y_min

    # Add text for number of observations and median values
    for i, category in enumerate(orderc):
        # Get counts for this category
        data_count = data_counts.get(category, 0)
        trash_count = trash_counts.get(category, 0)

        # Create tag for this category
        tag = f"{data_count:.0f}/{data_count + trash_count:.0f}"

        # Removed values tag
        ax.text(i, y_min - y_range * 0.07, tag,
                horizontalalignment='center', size='x-small', color='b', weight='semibold')

        # Median value tag - directly access the median for the current category
        if category in descriptor.index:
            median_val = round(descriptor.loc[category, "50%"], 2)
            ax.text(i, y_max + y_range * 0.05, str(median_val),
                    horizontalalignment='center', size='x-small', color='gray', weight='semibold')

    # Note about removed samples
    ax.text(1, 1, f"Removed: {trash[0].shape[0]}", horizontalalignment='right', verticalalignment='top',
            size='x-small', color='b', weight='semibold', transform=ax.transAxes)

    #plt.tight_layout()

    # save_excel_boxplot(path, data, var_x, var_full_y, var_y)
    wb = save_combined_excel_data(path, wb, data, trash, var_x, var_name_y, var_y, descriptor)

    if datatype == "junk":
        sample_name = f"boxplotj_{var_y}_by_{var_x}.png"
    else:
        sample_name = f"boxplot_{var_y}_by_{var_x}.png"

    if not is_running_in_jupyter():
        plt.savefig(f"{path}{sample_name}")
        plt.close()
        # plt.show()
        print(f"Saved boxplot of {var_y} by {var_x}")
    return fig, sample_name, wb


def is_running_in_jupyter():
    try:
        from IPython import get_ipython
        if 'IPKernelApp' not in get_ipython().config:  # Check if not in IPython kernel
            return False
    except Exception:
        return False
    return True


def save_full_data_frame(path, data):
    file_path = path + "0_numerical_results.xlsx"

    # Check if the Excel file already exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove the default sheet

    if not is_running_in_jupyter():
        with pd.ExcelWriter(file_path) as writer:
            # Write the DataFrame with earlier data to a sheet named 'Earlier'
            data.to_excel(writer, sheet_name='All_data')
    return wb


def save_combined_excel_data(path, wb, data, filtered_info, var_x, name_y, var_y, other_df):
    trash, filters = filtered_info
    # Define the Excel file path
    file_path = path + "0_numerical_results.xlsx"

    # # Check if the Excel file already exists
    # if os.path.exists(file_path):
    #     wb = load_workbook(file_path)
    # else:
    #     wb = openpyxl.Workbook()
    #     wb.remove(wb.active)  # Remove the default sheet

    # Create a new sheet name based on var_x and var_y
    sheet_title = f"{var_y}-by-{var_x}"

    # Check if the sheet already exists
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    # Insert personalized string before the first DataFrame
    ws.append([f"Contents of boxplot for {var_y} by {var_x}"])
    ws.append([])  # Add an empty row for spacing

    # Process and append data and other_df as before
    combined_data = data.copy()
    combined_data['_index'] = combined_data.groupby(var_x).cumcount()
    pivot_table = combined_data.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

    for r in dataframe_to_rows(pivot_table, index=True, header=True):
        ws.append(r)

    # Calculate starting row for the second personalized string
    # It's the current number of rows plus 2 for spacing
    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    ws.cell(row=next_row, column=1, value="Statistical summary")
    ws.append([])  # Add an empty row for spacing

    for r in dataframe_to_rows(other_df.T, index=True, header=True):
        ws.append(r)

    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    ws.cell(row=next_row, column=1, value="This is the filtered data")
    ws.append([])  # Add an empty row for spacing

    combined_trash = trash.copy()
    combined_trash['_index'] = combined_trash.groupby(var_x).cumcount()
    pivot_table_trash = combined_trash.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

    # Add rows from the second DataFrame (pivot table)
    for r in dataframe_to_rows(pivot_table_trash, index=True, header=True):
        ws.append(r)

    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    filter_words = ["Only data within these limits is shown:"] + filters
    for cc, strings in enumerate(filter_words):
        ws.cell(row=next_row + cc, column=1, value=strings)
    ws.append([])  # Add an empty row for spacing

    if not is_running_in_jupyter():
        # Save the workbook
        wb.save(filename=file_path)
    return wb


def histogram(path, df, var_y):
    names_dict = {
        'voc': 'Voc(V)', 'jsc': 'Jsc(mA/cm2)', 'ff': 'FF(%)', 'pce': 'PCE(%)',
        'vmpp': 'V_mpp(V)', 'jmpp': 'J_mpp(mA/cm2)', 'pmpp': 'P_mpp(mW/cm2)',
        'rser': 'R_series(Ohmcm2)', 'rshu': 'R_shunt(Ohmcm2)'
    }

    pl_y = names_dict[var_y]

    if var_y == "voc":
        bins = 20
    elif var_y == "jsc":
        bins = 30
    else:
        bins = 40

    plt.style.use('seaborn-v0_8-darkgrid')
    fig, ax = plt.subplots(figsize=(10, 6))  # Create figure and axes object

    # Use the axes object for the histogram
    df.hist(column=pl_y, bins=bins, grid=True, ax=ax)

    # Adjust the histogram's appearance if necessary
    # ax.set_title(f"Histogram of {pl_y}")
    # ax.set_xlabel(pl_y)
    # ax.set_ylabel("Frequency")

    sample_name = f"histogram_{var_y}.png"
    if not is_running_in_jupyter():
        # Save the figure
        fig.savefig(f"{path}/{sample_name}")
        print(f"Saved histogram of {var_y}")

    return fig, sample_name


def condition_string_test(condition_var, unique_values):
    # is_condition = False
    if len(condition_var) > 1:
        condition_list = condition_var.replace(" ", "").split(',')
        if len(condition_list) == len(unique_values):
            is_condition = True
            return is_condition, condition_list
        else:
            print(
                f"Provided conditions: {len(condition_list)}. Number of samples requiring conditions: "
                f"{len(unique_values)}.")
            condition_string = input(
                f"Please re-enter, providing exactly {len(unique_values)} conditions for the samples, "
                f"each separated by a comma: ")

            condition_string_test(condition_string, unique_values)
    else:
        is_condition = False
        return is_condition, []


def plot_list_from_voila(plot_list):
    jvc_dict = {'Voc': 'v', 'Jsc': 'j', 'FF': 'f', 'PCE': 'p', 'R_ser': 'r', 'R_shu': 'h', 'V_mpp': 'u',
                'i_mpp': 'p', 'P_mpp': 'm'}
    box_dict = {'by Batch': 'e', 'by Variable': 'g', 'by Sample': 'a', 'by Cell': 'b', 'by Scan Direction': 'c'}
    cur_dict = {'All cells': 'Cy', 'Only working cells': 'Cz', 'Only not working cells': 'Co', 'Best device only': 'Cw',
                'Separated by cell': 'Cx', 'Separated by substrate': 'Cd'}

    new_list = []
    for plot in plot_list:
        code = ''
        if "omitted" in plot[0]:
            code += "J"
            code += (jvc_dict[plot[1]])
            code += (box_dict[plot[2]])
        elif "Boxplot" in plot[0]:
            code += "B"
            code += (jvc_dict[plot[1]])
            code += (box_dict[plot[2]])
        elif "Histogram" in plot[0]:
            code += "H"
            code += (jvc_dict[plot[1]])
        else:
            code += (cur_dict[plot[1]])
        new_list.append(code)

    return new_list


def plotting_string_action(plot_list, wb, data, supp, is_voila=False):
    filtered_jv, complete_jv, complete_cur = data
    omitted_jv, filter_pars, is_conditions, path, samples = supp

    if is_voila:
        plot_list = plot_list_from_voila(plot_list)

    # varplot_dict = {"B": "boxplot", "H": "Histogram", "C": "JV curve"}
    varx_dict = {"a": "sample", "b": "cell", "c": "direction", "d": "ilum", "e": "batch", "g": "condition", }
    vary_dict = {"v": "voc", "j": "jsc", "f": "ff", "p": "pce", "u": "vmpp", "i": "jmpp", "m": "pmpp", "r": "rser",
                 "h": "rshu", }
    # varc_dict = {"w": "best device", "x": "all cells per sample", "y": "all together"}

    fig_list = []
    fig_names = []
    for pl in plot_list:
        # Check if there is "condition":
        if "g" in pl and not is_conditions:
            continue
        # Check and assign var_x
        for key, value in varx_dict.items():
            if key in pl:
                var_x = value
                break  # Found var_x, no need to check further
        else:  # No break occurred
            var_x = None

        # Check and assign var_y
        for key, value in vary_dict.items():
            if key in pl:
                var_y = value
                break  # Found var_y, no need to check further
        else:  # No break occurred
            var_y = None

        # Check and plot varplot
        if "B" in pl and var_x is not None and var_y is not None:
            print(wb, var_x, var_y)
            fig, fig_name, wb = boxplot_all_cells(path, wb, filtered_jv, var_x, var_y, [omitted_jv, filter_pars], "data")
        elif "J" in pl and var_x is not None and var_y is not None:
            fig, fig_name, wb = boxplot_all_cells(path, wb, omitted_jv, var_x, var_y, [filtered_jv, filter_pars], "junk")
        elif "H" in pl and var_y is not None:
            fig, fig_name = histogram(path, complete_jv, var_y)
        elif "Cw" in pl:  # Best device
            fig, fig_name = jv_plot_curve_best(path, complete_jv, complete_cur)
        elif "Cx" in pl:  # Cells per sample
            for s in samples:
                fig, fig_name = jv_plot_by_cell_3x2(complete_cur, s, path)
                fig_list.append(fig)
                fig_names.append(fig_name)
            continue
        elif "Cd" in pl:  # Cells per substrate
            for s in samples:
                fig, fig_name = jv_plot_by_substrate(complete_cur, s, path)
                fig_list.append(fig)
                fig_names.append(fig_name)
            continue
        elif "Cy" in pl:  # All data
            fig, fig_name = jv_plot_together(complete_jv, complete_cur, path, "All")
        elif "Cz" in pl:  # Only filtered
            fig, fig_name = jv_plot_together(filtered_jv, complete_cur, path, "Filtered")
        elif "Co" in pl:  # Only omitted
            fig, fig_name = jv_plot_together(omitted_jv, complete_cur, path, "Omitted")
        else:
            print(f"Command {pl} not recognized")
            continue

        fig_list.append(fig)
        fig_names.append(fig_name)

    return fig_list, fig_names, wb


def ask_to_input_initial_folder():
    user_path = input("Enter the path to the data folder for analysis. For example, C:\\Data\\Experiment : ")
    # user_path = r"D:\Seafile\JVData\Osail\20230717"
    path = user_path + '/'

    is_directory = os.path.isdir(path)

    if is_directory:
        listed_files = find_and_list_files(path)
        df_jvc, df_cur = load_files(listed_files)
        return df_jvc, df_cur, path
    else:
        print("Folder not found\n")
        return ask_to_input_initial_folder()


def find_unique_values(jvc_df):
    try:
        unique_values = jvc_df["identifier"].unique()
    except:
        unique_values = jvc_df["sample"].unique()
    print(f"\nThe following samples were found in the dataset: {', '.join(map(str, unique_values))}")

    return unique_values


def gather_conditions(unique_values):
    condition_var = input(
        "\nPress Enter to skip adding conditions. "
        "\nTo specify conditions for each sample, enter them no following the same order as above. "
        "Separate each condition with a comma. Leave a space for samples you wish to skip. "
        "Example: 1000 rpm, , 2500 rpm, 2500 rpm, 5000 rpm, ... :"
    )

    is_condition, list_conditions = condition_string_test(condition_var, unique_values)
    return is_condition, list_conditions


def gather_wanted_plots():
    plotting_string = input(
        "\nPress Enter to generate default plots: Boxplots (Voc, Jsc, FF, PCE) by sample, "
        "Histogram of PCE and all JV curves together."
        "\n\nTo create custom plots, enter codes as follows:"
        "\n  Plot Types: B=Boxplot, J=Boxplot(omitted), H=Histogram, C=JV curve"
        "\n  Parameters: a=sample, b=cell, c=direction, d=ilum, e=batch, g=condition"
        "\n              v=voc, j=jsc, f=ff, p=pce, u=vmpp, i=jmpp, m=pmpp, r=rser, h=rshu"
        "\n  JV Specific: Cw=best device only, Cx=only cells per sample, Cy=all data, Cz=only filtered, Co=only omitted"
        "\nExamples:"
        "\n  Bpg for a boxplot of PCEs by condition, Hv for a histogram of Voc,"
        "\n  Cy for all JV curves in a single plot."
        "\nWrite codes below separated by a comma, "
        "\nExample: Bfb, Hv, Cy : "
    )
    if len(plotting_string) > 1:
        plotting_list = plotting_string.replace(" ", "").split(',')
    else:
        plotting_list = ["Bav", "Baj", "Baf", "Bap", "Hp", "Cy"]

    return plotting_list


def create_new_results_folder(path):
    # Specify the path of the new folder
    folder_path = path + 'Results/'

    # Create the folder
    try:
        os.makedirs(folder_path)
    except FileExistsError:
        pass

    return folder_path


if __name__ == "__main__":
    # Individual actions
    jvc_data, cur_data, folder = ask_to_input_initial_folder()
    save_folder = create_new_results_folder(folder)
    workbook = save_full_data_frame(save_folder, jvc_data)

    if not jvc_data.empty:
        unique_vals = find_unique_values(jvc_data)
        is_cond, conditions = gather_conditions(unique_vals)
        jvc_data = name_by_condition(jvc_data, unique_vals, conditions)
        cur_data = name_by_condition(cur_data, unique_vals, conditions)
        jvc_filtered, junk, filter_vals = data_filter_setup(jvc_data, None)

        data_lists = [jvc_filtered, jvc_data, cur_data]
        extras = [junk, filter_vals, is_cond, save_folder, unique_vals]

        list_plots = gather_wanted_plots()
        # list_plots = ["Cz"]
        plotting_string_action(list_plots, workbook, data_lists, extras)

    print("Finished")