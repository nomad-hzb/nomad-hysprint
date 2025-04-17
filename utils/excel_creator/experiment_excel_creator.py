import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime

class ExperimentExcelBuilder:
    def __init__(
        self,
        process_sequence,
        is_testing=False
    ):
        """
        :param process_sequence: List of process dictionaries (e.g. [{'process': 'Spin Coating', 'config': {...}}, ...]).
        :param is_testing:       If True, row 3 of each column will display test values if steps are (label, test_value).
        """
        self.process_sequence = process_sequence

        self.workbook = Workbook()
        self.worksheet = self.workbook.active

        # Mark whether we want that third row for test values
        self.is_testing = is_testing

    @staticmethod
    def lighten_color(hex_color, factor=0.50):
        """
        Lightens a hex color by the given factor (0.50 = 50% lighter).
        """
        hex_color = hex_color.lstrip('#')
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)

        r = int(r + (255 - r) * factor)
        g = int(g + (255 - g) * factor)
        b = int(b + (255 - b) * factor)

        return f"{r:02x}{g:02x}{b:02x}".upper()

    def build_excel(self):
        start_col = 1
        incremental_number = 0  # Start from 1 for labeling processes

        for process_data in self.process_sequence:
            process_name = process_data['process']
            custom_config = process_data.get('config', {})

            # Pick a color from the global palette (defined outside the class)
            color_index = incremental_number % len(colors)
            cell_color = colors[color_index]

            # Generate the dynamic steps for this process
            steps = self.generate_steps_for_process(process_name, custom_config)

            step_count = len(steps)
            end_col = start_col + step_count - 1

            # # Label for the process (prefix with the incremental number, except for "Experiment Info")
            if process_name != "Experiment Info":
                process_label = f"{incremental_number}: {process_name}"
                # incremental_number += 1
            else:
                process_label = process_name

            # Merge the cells in Row 1 for the process name
            self.worksheet.merge_cells(
                start_row=1,
                start_column=start_col,
                end_row=1,
                end_column=end_col
            )

            # Row 1: Process label with the original color
            cell = self.worksheet.cell(row=1, column=start_col)
            cell.value = process_label
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")

            # Row 2 (and possibly Row 3 if is_testing=True) for steps
            row2_color = self.lighten_color(cell_color)  # Slightly lighter color for row 2
            row3_color = self.lighten_color(row2_color, 0.30)  # Even lighter for row 3 if needed

            for i, step_item in enumerate(steps):
                col_index = start_col + i

                # If the step is a tuple, interpret as (label, test_value)
                if isinstance(step_item, tuple):
                    step_label, test_val = step_item

                    # Row 2: the label
                    cell = self.worksheet.cell(row=2, column=col_index)
                    cell.value = step_label
                    cell.fill = PatternFill(start_color=row2_color, end_color=row2_color, fill_type="solid")

                    # Row 3: the test value (only if is_testing)
                    if self.is_testing:
                        for i in range(3, 4):
                            cell = self.worksheet.cell(row=i, column=col_index)
                            cell.value = test_val

            # Move to the next set of columns
            start_col = end_col + 1
            incremental_number += 1

        # Apply a custom formula for the "Nomad ID" column (example only)
        self.apply_nomad_id_formula()

        # Adjust the width of the columns based on their contents
        self.adjust_column_widths()

    def generate_steps_for_process(self, process_name, config):
        """
        This method constructs steps for each process. If is_testing=True,
        call make_label("Step Name", <test_value>) to pass a custom test value.
        Otherwise, just pass "Step Name".
        """

        # Helper to produce (label, test_val) if testing, or just label otherwise
        def make_label(label, test_val=None):
            if self.is_testing:
                # If no explicit test_val was provided, generate a default placeholder
                if test_val is not None:
                    return (label, test_val)
                else:
                    return (label, f"Test for {label}")
            else:
                return label

        if process_name == "Experiment Info":
            return [
                # Using this format to speed up testing with nomad
                # The format is: STEP, TEST_VARIABLE
                make_label("Date", "26-02-2025"),
                make_label("Project_Name", "FiNa"),
                make_label("Batch", "1"),
                make_label("Subbatch", "1"),
                make_label("Sample", "1"),
                make_label("Nomad ID", ""),
                make_label("Variation", "1000 rpm"),
                make_label("Sample dimension", "1 cm x 1 cm"),
                make_label("Sample area [cm^2]", 0.16),
                make_label("Number of pixels", 6),
                make_label("Pixel area [cm^2]", 0.16),
                make_label("Substrate material", "Soda Lime Glass"),
                make_label("Substrate conductive layer", "ITO"),
                make_label("Number of junctions", 1),
                make_label("Notes", "Test excel")
            ]

        if process_name == "Cleaning O2-Plasma" or process_name == "Cleaning UV-Ozone":
            steps = []
            for i in range(1, config.get("solvents", 0) + 1):
                steps.extend([
                    make_label(f"Solvent {i}", f"Hellmanex"),
                    make_label(f"Time {i} [s]", 30 + i),
                    make_label(f"Temperature {i} [°C]", 60 + i)
                ])

            if process_name == "Cleaning O2-Plasma":
                steps.extend([
                    make_label("Gas-Plasma Gas", "Oxygen"),
                    make_label("Gas-Plasma Time [s]", 180),
                    make_label("Gas-Plasma Power [W]", 50)
                ])
            if process_name == "Cleaning UV-Ozone":
                steps.append(make_label("UV-Ozone Time [s]", 900))
            return steps

        if process_name in ["Spin Coating", "Dip Coating", "Slot Die Coating", "Inkjet Printing"]:
            steps = [
                make_label("Material name", f"Cs0.05(MA0.17FA0.83)0.95Pb(I0.83Br0.17)3"),
                make_label("Layer type", f"Absorber"),
                make_label("Tool/GB name", f"HZB-HySprintBox")
            ]

            # Add solvent steps
            for i in range(1, config.get('solvents', 0) + 1):
                steps.extend([
                    make_label(f"Solvent {i} name", f"DMF"),
                    make_label(f"Solvent {i} volume [uL]", 10 * i),
                    make_label(f"Solvent {i} relative amount", 1.5)
                ])

            # Add solute steps
            for i in range(1, config.get('solutes', 0) + 1):
                steps.extend([
                    make_label(f"Solute {i} name", f"PbI2"),
                    make_label(f"Solute {i} Concentration [mM]", 1.42)
                ])

            # Add process-specific steps
            if process_name == "Spin Coating":
                steps.extend([
                    make_label("Solution volume [um]", 100),
                    make_label("Spin Delay [s]", 0.5)
                ])

                if config.get('spinsteps', 0) == 1:
                    steps.extend([
                        make_label("Rotation speed [rpm]", 1500),
                        make_label("Rotation time [s]", 30),
                        make_label("Acceleration [rpm/s]", 500)
                    ])
                else:
                    for i in range(1, config.get('spinsteps', 0) + 1):
                        steps.extend([
                            make_label(f"Rotation speed {i} [rpm]", 3000 + i),
                            make_label(f"Rotation time {i} [s]", 30 + i),
                            make_label(f"Acceleration {i} [rpm/s]", 1000 + i)
                        ])

                if config.get("antisolvent", False):
                    steps.extend([
                        make_label("Anti solvent name", "Toluene"),
                        make_label("Anti solvent volume [ml]", 0.3),
                        make_label("Anti solvent dropping time [s]", 25),
                        make_label("Anti solvent dropping speed [ul/s]", 50),
                        make_label("Anti solvent dropping heigt [mm]", 30)
                    ])

                if config.get("gasquenching", False):
                    steps.extend([
                        make_label("Gas", "Nitrogen"),
                        make_label("Gas quenching start time [s]", 5),
                        make_label("Gas quenching duration [s]", 15),
                        make_label("Gas quenching flow rate [ml/s]", 20),
                        make_label("Gas quenching pressure [bar]", 1.2),
                        make_label("Gas quenching velocity [m/s]", 2.5),
                        make_label("Gas quenching height [mm]", 10),
                        make_label("Nozzle shape", "Round"),
                        make_label("Nozzle size [mm²]", 3)
                    ])

                if config.get("vacuumquenching", False):
                    steps.extend([
                        make_label("Vacuum quenching start time [s]", 8),
                        make_label("Vacuum quenching duration [s]", 20),
                        make_label("Vacuum quenching pressure [bar]", 0.01)
                    ])

            elif process_name == "Slot Die Coating":
                steps.extend([
                    make_label("Solution volume [um]", 100),
                    make_label("Flow rate [ul/min]", 25),
                    make_label("Head gap [mm]", 0.3),
                    make_label("Speed [mm/s]", 15),
                    make_label("Air knife angle [°]", 45),
                    make_label("Air knife gap [cm]", 0.5),
                    make_label("Bead volume [mm/s]", 2),
                    make_label("Drying speed [cm/min]", 30)
                ])

            elif process_name == "Dip Coating":
                steps.append(make_label("Dipping duration [s]", 15))

            elif process_name == "Inkjet Printing":
                steps.extend([
                    make_label("Printhead name", "Spectra 0.8uL"),
                    make_label("Number of active nozzles", 128),
                    make_label("Droplet density [dpi]", 400),
                    make_label("Quality factor", 3),
                    make_label("Step size", 10),
                    make_label("Printed area [mm²]", 100),
                    make_label("Droplet per second [1/s]", 5000),
                    make_label("Droplet volume [pl]", 10),
                    make_label("Ink reservoir pressure [bar]", 0.3),
                    make_label("Table temperature [°C]", 40),
                    make_label("Nozzle temperature [°C]", 35),
                    make_label("rel. humidity [%]", 45)
                ])

            # Add annealing steps for all coating processes
            steps.extend([
                make_label("Annealing time [min]", 30),
                make_label("Annealing temperature [°C]", 120),
                make_label("Annealing atmosphere", "Nitrogen"),
                make_label("Notes", "Test annealing")
            ])

            return steps

        # PVD Processes
        if process_name == "Evaporation" or process_name == "Sublimation":
            steps = [
                make_label("Material name", f"PCBM"),
                make_label("Layer type", "Electron Transport Layer"),
                make_label("Tool/GB name", f"Hysprint Evap"),
                make_label("Organic", True),
                make_label("Base pressure [bar]", 1e-6),
                make_label("Pressure start [bar]", 5e-6),
                make_label("Pressure end [bar]", 3e-6),
                make_label("Source temperature start[°C]", 150),
                make_label("Source temperature end[°C]", 160),
                make_label("Substrate temperature [°C]", 25),
                make_label("Thickness [nm]", 100),
                make_label("Rate start [angstrom/s]", 0.5),
                make_label("Rate target [angstrom/s]", 1.0),
                make_label("Tooling factor", 1.5),
                make_label("Notes", f"Test note")
            ]
            return steps

        if process_name == "Seq-Evaporation" or process_name == "Co-Evaporation":
            steps = [
                make_label("Material name", f"Aluminium"),
                make_label("Layer type", "Electrode"),
                make_label("Tool/GB name", f"IRIS Evap"),
                make_label("Base pressure [bar]", 1e-6),
                make_label("Pressure start [bar]", 5e-6),
                make_label("Pressure end [bar]", 3e-6),
                make_label("Substrate temperature [°C]", 25)
            ]
            for i in range(1, config.get('materials', 0) + 1):
                steps.extend([
                    make_label(f"Material {i} name", f"Cupper"),
                    make_label(f"Material {i} organic", False),
                    make_label(f"Material {i} source temperature start [°C]", 100 + 10 + i),
                    make_label(f"Material {i} source temperature end [°C]", 110 + 10 + i),
                    make_label(f"Material {i} thickness [nm]", 20 + i),
                    make_label(f"Material {i} rate [angstrom/s]", 0.5 + i),
                    make_label(f"Material {i} tooling factor", 1.0 + 0.1 + i),
                    make_label("Notes", f"Test note co-evaporation")
                ])
            return steps

        if process_name == "Sputtering":
            steps = [
                make_label("Material name", "TiO2"),
                make_label("Layer type", "Electron Transport Layer"),
                make_label("Tool/GB name", "Hysprint tool"),
                make_label("Gas", "Argon"),
                make_label("Temperature [°C]", 200),
                make_label("Pressure [mbar]", 0.01),
                make_label("Deposition time [s]", 300),
                make_label("Burn in time [s]", 60),
                make_label("Power [W]", 150),
                make_label("Rotation rate [rpm]", 30),
                make_label("Thickness [nm]", 50),
                make_label("Gas flow rate [cm^3/min]", 20),
                make_label("Notes", "Test Sputtering")
            ]
            return steps

        if process_name == "Laser Scribing":
            steps = [
                make_label("Laser wavelength [nm]", 532),
                make_label("Laser pulse time [ps]", 8),
                make_label("Laser pulse frequency [kHz]", 80),
                make_label("Speed [mm/s]", 100),
                make_label("Fluence [J/cm2]", 0.5),
                make_label("Power [%]", 75),
                make_label("Recipe file", "test_scribing_recipe.xml")
            ]
            return steps

        if process_name == "ALD":
            steps = [
                make_label("Material name", "Al2O3"),
                make_label("Layer type", "Electron Transport Layer"),
                make_label("Tool/GB name", "IRIS ALD"),
                make_label("Source", "TMA"),
                make_label("Thickness [nm]", 25),
                make_label("Temperature [°C]", 150),
                make_label("Rate [A/s]", 0.1),
                make_label("Time [s]", 1800),
                make_label("Number of cycles", 250),
                make_label("Precursor 1", "TMA"),
                make_label("Pulse duration 1 [s]", 0.2),
                make_label("Manifold temperature 1 [°C]", 80),
                make_label("Bottle temperature 1 [°C]", 25),
                make_label("Precursor 2 (Oxidizer/Reducer)", "H2O"),
                make_label("Pulse duration 2 [s]", 0.1),
                make_label("Manifold temperature 2 [°C]", 70)
            ]
            return steps

        if process_name == "Annealing":
            steps = [
                make_label("Annealing time [min]", 60),
                make_label("Annealing temperature [°C]", 150),
                make_label("Annealing athmosphere", "Nitrogen"),
                make_label("Relative humidity [%]", 35),
                make_label("Notes", "Test annealing process")
            ]
            return steps

        if process_name == "Generic Process":
            steps = [
                make_label("Name", "Test Generic Process"),
                make_label("Notes", "This is a test generic process")
            ]
            return steps

        else:
            print(f"Warning: Process '{process_name}' not defined in generate_steps_for_process. Using default steps.")
            return [make_label("Undefined Process", "Test value")]

    def adjust_column_widths(self):
        for col in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # Get the column letter
            for cell in col:
                if cell.value and isinstance(cell.value, str):
                    max_length = max(max_length, len(cell.value))
            adjusted_width = (max_length + 2)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width

    def apply_nomad_id_formula(self):
        for row in range(3, 4):  # Adjusted range for testing
            nomad_id_formula = f'=CONCATENATE("HZB_",B{row},"_",C{row},"_",D{row},"_C-",E{row})'
            self.worksheet[f"F{row}"].value = nomad_id_formula

    def save(self):
        current_date = datetime.now().strftime("%Y%m%d")
        filename = f"{current_date}_experiment_file.xlsx"
        self.workbook.save(filename)
        print(f"File saved as: {filename}")


# Define Tableau Colors
TABLEAU_COLORS = {
    'tab:blue': '1F77B4',
    'tab:orange': 'FF7F0E',
    'tab:green': '2CA02C',
    'tab:red': 'D62728',
    'tab:purple': '9467BD',
    'tab:brown': '8C564B',
    'tab:pink': 'E377C2',
    'tab:gray': '7F7F7F',
    'tab:olive': 'BCBD22',
    'tab:cyan': '17BECF'
}

# Extract just the hex values
colors = list(TABLEAU_COLORS.values())

# Define the sequence of processes with custom configurations

# # Test process with all processes
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Cleaning O2-Plasma", "config": {"solvents": 4}},
#     {"process": "Cleaning UV-Ozone", "config": {"solvents": 2}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Dip Coating", "config":  {"solvents": 2,
#                                             "solutes": 5, "spinsteps": 2, "antisolvent": True}},
#     {"process": "Slot Die Coating", "config":  {"solvents": 1, "solutes": 2,
#                                             "spinsteps": 1, "antisolvent": False}},
#     {"process": "Inkjet Printing", "config":  {"solvents": 1, "solutes": 2,
#                                             "spinsteps": 1, "antisolvent": False}},
#     {"process": "Evaporation"},
#     {"process": "Sublimation"},
#     {"process": "Seq-Evaporation"},
#     {"process": "Co-Evaporation"},
#     {"process": "Sputtering"},
#     {"process": "Laser Scribing"},
#     {"process": "ALD"},
#     {"process": "Annealing"},
#     {"process": "Generic Process"},
# ]

# # Micha's test
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Cleaning O2-Plasma", "config": {"solvents": 1}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": True}},
#     {"process": "Inkjet Printing", "config":  {"solvents": 1, "solutes": 2,
#                                             "spinsteps": 1, "antisolvent": False}},
#     {"process": "ALD"},
#     {"process": "Laser Scribing"},
#      {"process": "Slot Die Coating", "config":  {"solvents": 1, "solutes": 2,
#                                              "spinsteps": 1, "antisolvent": False}},
#     {"process": "Evaporation"},
#     {"process": "Generic Process"},

# # Osail's bypass
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Cleaning O2-Plasma", "config": {"solvents": 4}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # PEDOT-PSS
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Evaporation"},  # BC60
#     {"process": "Evaporation"},  # BCP
#     {"process": "Evaporation"},  # Aluminium
#     ]

# # Anne's bypass
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Cleaning UV-Ozone", "config": {"solvents": 3}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # PEDOT-PSS or SAM
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 2,  # Perovskite
#                                             "solutes": 5, "spinsteps": 1, "antisolvent": True}},
#     {"process": "Evaporation"},  # BC60
#     {"process": "Evaporation"},  # BCP
#     {"process": "Evaporation"},  # Cupper
#     ]

# # Printer process
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Cleaning UV-Ozone", "config": {"solvents": 4}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # PEDOT-PSS or SAM
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Inkjet Printing", "config":  {"solvents": 3,  # Perovskite
#                                             "solutes": 5, "antisolvent": True}},
#     {"process": "Evaporation"},  # BC60
#     {"process": "Evaporation"},  # BCP
#     {"process": "Evaporation"},  # Cupper
#     ]

# Eike process
process_sequence = [
    {"process": "Experiment Info"},
    {"process": "Cleaning UV-Ozone", "config": {"solvents": 1}},
    {"process": "Spin Coating", "config":  {"solvents": 1,  # PEDOT-PSS or SAM
                                            "solutes": 1, "spinsteps": 2, "antisolvent": False}},
    {"process": "Spin Coating", "config": {"solvents": 1,  # Washing
                                           "solutes": 1, "spinsteps": 2, "antisolvent": False}},
    {"process": "Spin Coating", "config": {"solvents": 2,  # Perovskite
                                           "solutes": 6, "spinsteps": 2, "antisolvent": False}},
    {"process": "Spin Coating", "config": {"solvents": 1,  # Interlayer
                                           "solutes": 1, "spinsteps": 1, "antisolvent": False}},
    {"process": "Evaporation"},  # C60
    {"process": "ALD"},  # SnO2
    {"process": "Sputtering"},  # TCO
    {"process": "Evaporation"},  # BCP
    {"process": "Evaporation"},  # Cupper
    ]

# # Kevin Prince
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Cleaning UV-Ozone", "config": {"solvents": 4}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # NiOx
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # SAM
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 2,  # all-inorganic PK
#                                             "solutes": 8, "spinsteps": 2, "antisolvent": True}},
#     {"process": "Spin Coating", "config":  {"solvents": 2,  # PEAI Treatment
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # PCBM
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # PEIE
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "ALD"},  # SnOx
#     {"process": "Sputtering"},  # SnOx
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # SAM
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 2,  # all-inorganic PK
#                                             "solutes": 4, "spinsteps": 2, "antisolvent": True}},
#     {"process": "Spin Coating", "config": {"solvents": 1,  # treatment
#                                            "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Evaporation"},  # C60
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # PEIE
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "ALD"},  # SnOx
#     {"process": "Sputtering"},  # ITO
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # pedot:pss
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 2, # narrow PK
#                                             "solutes": 8, "spinsteps": 2, "antisolvent": True}},
#     {"process": "Spin Coating", "config":  {"solvents": 2,  # Treatment
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Evaporation"},  # C60
#     {"process": "ALD"},  # SnOx
#     {"process": "Evaporation"},  # Cu
# ]

# Process for Daniel Spinbot
# process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Spin Coating", "config":  {"solvents": 1,
#                                            "solutes": 1, "spinsteps": 1, "antisolvent": False}},  # SAM
#    {"process": "Spin Coating", "config":  {"solvents": 2,
#                                            "solutes": 5, "spinsteps": 2, "antisolvent": True}},  # PSK
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2,
#                                            "spinsteps": 1, "antisolvent": False}},  # Passivation Sol
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "ALD"},  # SnO2
#    {"process": "Evaporation"}  # Ag
# ]


# Process for Spinbot gasquenched
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#     {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps": 1}},  # SAM
#     {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 5, "spinsteps": 2, "gasquenching": True}},  # PSK
#     {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2, "spinsteps": 1}},  # Passivation Sol
#     {"process": "Evaporation"},  # Passivation Evap
#     {"process": "Evaporation"},  # C60
#     {"process": "Evaporation"},  # BCP
#     {"process": "Evaporation"}   # Ag
# ]

# process_sequence = [
#     # Philippe process
#     {"process": "Experiment Info"},
#     {"process": "Cleaning UV-Ozone", "config": {"solvents": 4}}, #FTO
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # NiOx
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # SAM
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 2,  # Perovkiste
#                                             "solutes": 4, "spinsteps": 2, "antisolvent": True}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # PCBM
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Spin Coating", "config":  {"solvents": 1,  # BCP
#                                             "solutes": 1, "spinsteps": 1, "antisolvent": False}},
#     {"process": "Evaporation"},
# ]

# Process for Hybrid
# Process_sequence = [
#   {"process": "Experiment Info"},
#   {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}},   #SAM
#   {"process": "Seq-Evaportation", "config":  {"materials": 2} },                                                 #PSK inorganic
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 3, "spinsteps":1 , "antisolvent": False}},   #PSK organic
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2, "spinsteps":1 , "antisolvent": False}},   #Passivation Sol
#   {"process": "Evaporation"},    #Passivation Evap
#   {"process": "Evaporation"},    #C60
#   {"process": "Evaporation"},    #BCP
#   {"process": "ALD"},            #SnO2
#   {"process": "Evaporation"}     #Ag
# ]


# Process for SOP
# process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 3}},
#    {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #NiO
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #SAM
#    {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 6, "spinsteps":2 , "antisolvent": True}}, #PSK
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #PEAI
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #PCBM
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #BCP
#    {"process": "Evaporation"} #Ag
# ]


# Test Process
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1}},   #SAM
#     {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 5, "spinsteps":2 , "antisolvent": True}},    #PSK
#     {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "gasquenching": True}},   #Passivation Sol
# ]

# Create an instance of ExperimentExcelBuilder and build the Excel file
#builder = ExperimentExcelBuilder(process_sequence, process_config, is_testing=True)
builder = ExperimentExcelBuilder(process_sequence, is_testing=True)

builder.build_excel()
builder.save()
