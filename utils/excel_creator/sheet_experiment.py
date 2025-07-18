from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

TABLEAU_COLORS = {
    'tab:blue': '1F77B4',
    'tab:orange': 'FF7F0E',
    'tab:green': '2CA02C',
    'tab:red': 'D62728',
    'tab:purple': '9467BD',
    'tab:brown': '8C564B',
    'tab:pink': 'E377C2',
    'tab:gray': '7F7F7F',
    'tab:olive': 'BCBD22',
    'tab:cyan': '17BECF',
}
colors = list(TABLEAU_COLORS.values())


def lighten_color(hex_color, factor=0.50):
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f'{r:02x}{g:02x}{b:02x}'.upper()


def add_experiment_sheet(workbook, process_sequence, is_testing=False):
    ws = workbook.active
    ws.title = 'Experiment Data'
    start_col = 1
    incremental_number = 0

    def make_label(label, test_val=None):
        if is_testing:
            if test_val is not None:
                return (label, test_val)
            else:
                return (label, f'Test for {label}')
        else:
            return label

    def generate_steps_for_process(process_name, config):
        """
        This method constructs steps for each process. If is_testing=True,
        call make_label("Step Name", <test_value>) to pass a custom test value.
        Otherwise, just pass "Step Name".
        """

        if process_name == 'Experiment Info':
            # Check if Ink Recycling exists in the sequence
            has_ink_recycling = any(p.get('process') == 'Ink Recycling' for p in process_sequence)

            if has_ink_recycling:
                # Simplified fields for ink recycling
                return [
                    make_label('Date', '27-05-2025'),
                    make_label('Project_Name', 'FiNa'),
                    make_label('Batch', '1'),
                    make_label('Subbatch', '1'),
                    make_label('Sample', '1'),
                    make_label('Nomad ID', ''),
                    make_label('Variation', 'Some variation'),
                ]
            
            # Original experiment info fields for other cases
            return [
                # Using this format to speed up testing with nomad
                # The format is: STEP, TEST_VARIABLE
                make_label('Date', '26-02-2025'),
                make_label('Project_Name', 'FiNa'),
                make_label('Batch', '1'),
                make_label('Subbatch', '1'),
                make_label('Sample', '1'),
                make_label('Nomad ID', ''),
                make_label('Variation', '1000 rpm'),
                make_label('Sample dimension', '1 cm x 1 cm'),
                make_label('Sample area [cm^2]', 0.16),
                make_label('Number of pixels', 6),
                make_label('Pixel area [cm^2]', 0.16),
                make_label('Substrate material', 'Soda Lime Glass'),
                make_label('Substrate conductive layer', 'ITO'),
                make_label('Number of junctions', 1),
                make_label('Notes', 'Test excel'),
            ]

        if process_name == 'Cleaning O2-Plasma' or process_name == 'Cleaning UV-Ozone':
            steps = []
            for i in range(1, config.get('solvents', 0) + 1):
                steps.extend(
                    [
                        make_label(f'Solvent {i}', 'Hellmanex'),
                        make_label(f'Time {i} [s]', 30 + i),
                        make_label(f'Temperature {i} [°C]', 60 + i),
                    ]
                )

            if process_name == 'Cleaning O2-Plasma':
                steps.extend(
                    [
                        make_label('Gas-Plasma Gas', 'Oxygen'),
                        make_label('Gas-Plasma Time [s]', 180),
                        make_label('Gas-Plasma Power [W]', 50),
                    ]
                )
            if process_name == 'Cleaning UV-Ozone':
                steps.append(make_label('UV-Ozone Time [s]', 900))
            return steps

        if process_name in ['Spin Coating', 'Dip Coating', 'Slot Die Coating', 'Inkjet Printing']:
            steps = [
                make_label('Material name', 'Cs0.05(MA0.17FA0.83)0.95Pb(I0.83Br0.17)3'),
                make_label('Layer type', 'Absorber'),
                make_label('Tool/GB name', 'HZB-HySprintBox'),
            ]

            # Add solvent steps
            for i in range(1, config.get('solvents', 0) + 1):
                steps.extend(
                    [
                        make_label(f'Solvent {i} name', 'DMF'),
                        make_label(f'Solvent {i} volume [uL]', 10 * i),
                        make_label(f'Solvent {i} relative amount', 1.5),
                        make_label(f'Solvent {i} chemical ID', '1592-461-04-2'),
                    ]
                )

            # Add solute steps
            for i in range(1, config.get('solutes', 0) + 1):
                steps.extend(
                    [
                        make_label(f'Solute {i} name', 'PbI2'),
                        make_label(f'Solute {i} Concentration [mM]', 1.42),
                        make_label(f'Solute {i} chemical ID', '2393-752-02-3'),
                    ]
                )

            # Add process-specific steps
            if process_name == 'Spin Coating':
                steps.extend([make_label('Solution volume [uL]', 100),
                             make_label('Spin Delay [s]', 0.5)])

                if config.get('spinsteps', 0) == 1:
                    steps.extend(
                        [
                            make_label('Rotation speed [rpm]', 1500),
                            make_label('Rotation time [s]', 30),
                            make_label('Acceleration [rpm/s]', 500),
                        ]
                    )
                else:
                    for i in range(1, config.get('spinsteps', 0) + 1):
                        steps.extend(
                            [
                                make_label(f'Rotation speed {i} [rpm]', 3000 + i),
                                make_label(f'Rotation time {i} [s]', 30 + i),
                                make_label(f'Acceleration {i} [rpm/s]', 1000 + i),
                            ]
                        )

                if config.get('antisolvent', False):
                    steps.extend(
                        [
                            make_label('Anti solvent name', 'Toluene'),
                            make_label('Anti solvent volume [ml]', 0.3),
                            make_label('Anti solvent dropping time [s]', 25),
                            make_label('Anti solvent dropping speed [ul/s]', 50),
                            make_label('Anti solvent dropping heigt [mm]', 30),
                        ]
                    )

                if config.get('gasquenching', False):
                    steps.extend(
                        [
                            make_label('Gas', 'Nitrogen'),
                            make_label('Gas quenching start time [s]', 5),
                            make_label('Gas quenching duration [s]', 15),
                            make_label('Gas quenching flow rate [ml/s]', 20),
                            make_label('Gas quenching pressure [bar]', 1.2),
                            make_label('Gas quenching velocity [m/s]', 2.5),
                            make_label('Gas quenching height [mm]', 10),
                            make_label('Nozzle shape', 'Round'),
                            make_label('Nozzle size [mm²]', 3),
                        ]
                    )

                if config.get('vacuumquenching', False):
                    steps.extend(
                        [
                            make_label('Vacuum quenching start time [s]', 8),
                            make_label('Vacuum quenching duration [s]', 20),
                            make_label('Vacuum quenching pressure [bar]', 0.01),
                        ]
                    )

            elif process_name == 'Slot Die Coating':
                steps.extend(
                    [
                        make_label('Solution volume [uL]', 100),
                        make_label('Flow rate [ul/min]', 25),
                        make_label('Head gap [mm]', 0.3),
                        make_label('Speed [mm/s]', 15),
                        make_label('Air knife angle [°]', 45),
                        make_label('Air knife gap [cm]', 0.5),
                        make_label('Bead volume [mm/s]', 2),
                        make_label('Drying speed [cm/min]', 30),
                        make_label('Chuck heating temperature [°C]', 25),
                    ]
                )

            elif process_name == 'Dip Coating':
                steps.append(make_label('Dipping duration [s]', 15))

            elif process_name == 'Inkjet Printing':
                steps.extend(
                    [
                        make_label('Printhead name', 'Spectra 0.8uL'),
                        make_label('Number of active nozzles', 128),
                        make_label('Active nozzles', "all"),
                        make_label('Droplet density X [dpi]', 400),
                        make_label('Droplet density Y [dpi]', 300),
                        make_label('Quality factor', 3),
                        make_label('Step size', 10),
                        make_label('Printing direction', 10),
                        make_label('Number of swaths', 10),
                        make_label('Printed area [mm²]', 100),
                        make_label('Droplet per second [1/s]', 5000),
                        make_label('Droplet volume [pl]', 10),
                        make_label('Ink reservoir pressure [bar]', 0.3),
                        make_label('Table temperature [°C]', 40),
                        make_label('Dropping Height [mm]', 12),
                        make_label('Substrate thickness [mm]', 20),
                        make_label('Printing speed [mm/s]', 10),
                        make_label('Print head angle [deg]', 13),
                        make_label('Nozzle temperature [°C]', 35),
                        make_label('Nozzle voltage config file', "testfile.txt"),
                        make_label('Image used', "Square inch 300 dpi"),
                        make_label('rel. humidity [%]', 45),
                    ]
                )

                if config.get('gavd', False):
                    steps.extend(
                        [
                            make_label('GAVD Gas', 'Nitrogen'),
                            make_label('GAVD start time [s]', 5),
                            make_label('GAVD vacuum pressure [mbar]', 10),
                            make_label('GAVD temperature [°C]', 25),
                            make_label('GAVD vacuum time [s]', 15),
                            make_label('Gas flow duration [s]', 15),
                            make_label('Gas flow pressure [mbar]', 100),
                            make_label('Nozzle shape', 'round'),
                            make_label('Nozzle type', 'mesh'),
                            make_label('GAVD comment', 'blabla'),
                        ]
                    )

            # Add annealing steps for all coating processes
            steps.extend(
                [
                    make_label('Annealing time [min]', 30),
                    make_label('Annealing temperature [°C]', 120),
                    make_label('Annealing atmosphere', 'Nitrogen'),
                    make_label('Notes', 'Test annealing'),
                ]
            )

            return steps

        # PVD Processes
        if process_name == 'Evaporation' or process_name == 'Sublimation':
            steps = [
                make_label('Material name', 'PCBM'),
                make_label('Layer type', 'Electron Transport Layer'),
                make_label('Tool/GB name', 'Hysprint Evap'),
                make_label('Organic', True),
                make_label('Base pressure [bar]', 1e-6),
                make_label('Pressure start [bar]', 5e-6),
                make_label('Pressure end [bar]', 3e-6),
                make_label('Source temperature start[°C]', 150),
                make_label('Source temperature end[°C]', 160),
                make_label('Substrate temperature [°C]', 25),
                make_label('Thickness [nm]', 100),
                make_label('Rate start [angstrom/s]', 0.5),
                make_label('Rate target [angstrom/s]', 1.0),
                make_label('Tooling factor', 1.5),
                make_label('Notes', 'Test note'),
            ]
            return steps

        # if process_name == 'Seq-Evaporation' or process_name == 'Co-Evaporation':
        if process_name == 'Co-Evaporation':
            steps = [
                make_label('Material name', 'Aluminium'),
                make_label('Layer type', 'Electrode'),
                make_label('Tool/GB name', 'IRIS Evap'),
            ]
            for i in range(1, config.get('materials', 0) + 1):
                steps.extend(
                    [
                        make_label(f'Material name {i}', 'Cupper'),
                        make_label(
                            f'Source temperature start {i}[°C]', 100 + 10 + i),
                        make_label(
                            f'Source temperature end {i}[°C]', 110 + 10 + i),
                        make_label(f'Thickness {i} [nm]', 20 + i),
                        make_label(f'Rate {i} [angstrom/s]', 0.5 + i),
                        make_label(f'Base pressure {i} [bar]', 1e-6),
                        make_label(f'Pressure start {i} [bar]', 5e-6),
                        make_label(f'Pressure end {i} [bar]', 3e-6),
                        make_label(f'Substrate temperature {i} [°C]', 25),
                        make_label(f'Tooling factor {i}', 1.0 + 0.1 + i)
                    ]
                )
            steps.append(make_label('Notes', 'Test note co-evaporation'))
            return steps

        if process_name == 'Sputtering':
            steps = [
                make_label('Material name', 'TiO2'),
                make_label('Layer type', 'Electron Transport Layer'),
                make_label('Tool/GB name', 'Hysprint tool'),
                make_label('Gas', 'Argon'),
                make_label('Temperature [°C]', 200),
                make_label('Pressure [mbar]', 0.01),
                make_label('Deposition time [s]', 300),
                make_label('Burn in time [s]', 60),
                make_label('Power [W]', 150),
                make_label('Rotation rate [rpm]', 30),
                make_label('Thickness [nm]', 50),
                make_label('Gas flow rate [cm^3/min]', 20),
                make_label('Notes', 'Test Sputtering'),
            ]
            return steps

        if process_name == 'Laser Scribing':
            steps = [
                make_label('Laser wavelength [nm]', 532),
                make_label('Laser pulse time [ps]', 8),
                make_label('Laser pulse frequency [kHz]', 80),
                make_label('Speed [mm/s]', 100),
                make_label('Fluence [J/cm2]', 0.5),
                make_label('Power [%]', 75),
                make_label('Recipe file', 'test_scribing_recipe.xml'),
                make_label('Dead area [cm2]', 2),
                make_label('Width of cell [mm]', 5),
                make_label('Number of cells', 6),
                make_label('Notes', "Laser Note"),
            ]
            return steps

        if process_name == 'ALD':
            steps = [
                make_label('Material name', 'Al2O3'),
                make_label('Layer type', 'Electron Transport Layer'),
                make_label('Tool/GB name', 'IRIS ALD'),
                make_label('Source', 'TMA'),
                make_label('Thickness [nm]', 25),
                make_label('Temperature [°C]', 150),
                make_label('Rate [A/s]', 0.1),
                make_label('Time [s]', 1800),
                make_label('Number of cycles', 250),
                make_label('Precursor 1', 'TMA'),
                make_label('Pulse duration 1 [s]', 0.2),
                make_label('Manifold temperature 1 [°C]', 80),
                make_label('Bottle temperature 1 [°C]', 25),
                make_label('Precursor 2 (Oxidizer/Reducer)', 'H2O'),
                make_label('Pulse duration 2 [s]', 0.1),
                make_label('Manifold temperature 2 [°C]', 70),
            ]
            return steps

        if process_name == 'Annealing':
            steps = [
                make_label('Annealing time [min]', 60),
                make_label('Annealing temperature [°C]', 150),
                make_label('Annealing athmosphere', 'Nitrogen'),
                make_label('Relative humidity [%]', 35),
                make_label('Notes', 'Test annealing process'),
            ]
            return steps

        if process_name == 'Generic Process':
            steps = [
                make_label('Name', 'Test Generic Process'),
                make_label('Notes', 'This is a test generic process'),
            ]
            return steps

        if process_name == 'Ink Recycling':
            steps = []
            # Ink Preparation steps
            for i in range(1, config.get('solvents', 0) + 1):
                steps.extend([
                    make_label(f'Solvent {i} name', f'DMF {i}'),
                    make_label(f'Solvent {i} volume [ml]', 10 * i),
                ])
            for i in range(1, config.get('solutes', 0) + 1):
                steps.extend([
                    make_label(f'Solute {i} name', f'PbI2 {i}'),
                    make_label(f'Solute {i} concentration [M]', 1.5 * i),
                    make_label(f'Solute {i} amount [g]', 5.0 * i),
                    make_label(f'Solute {i} moles [mol]', 0.02 * i),
                ])
            for i in range(1, config.get('precursors', 0) + 1):
                steps.extend([
                    make_label(f'Precursor {i} name', f'MAI {i}'),
                    make_label(f'Precursor {i} moles [mol]', 0.01 * i),
                ])
            
            # Mixing steps
            steps.extend([
            make_label('Functional liquid name', 'FL'),
            make_label('Functional liquid volume [ml]', 25),
            make_label('Dissolving temperature [°C]', 60),
            ])

            # Filtering steps
            steps.extend([
            make_label('Filter material', 'Paper'),
            make_label('Filter size [mm]', 0.45),
            make_label('Filter weight [g]', 0.5),
            ])

            # Results steps
            steps.extend([
            make_label('Recovered solute [g]', 4.2),
            make_label('Yield [%]', 84),
            make_label('Notes', 'Test recycling process'),
            ])
            return steps

        else:
            print(
            f"Warning: Process '{process_name}' not defined in generate_steps_for_process. Using default steps."
            )
            return [make_label('Undefined Process', 'Test value')]

    for process_data in process_sequence:
        process_name = process_data['process']
        custom_config = process_data.get('config', {})
        color_index = incremental_number % len(colors)
        cell_color = colors[color_index]
        steps = generate_steps_for_process(process_name, custom_config)
        step_count = len(steps)
        end_col = start_col + step_count - 1

        if process_name != 'Experiment Info':
            process_label = f'{incremental_number}: {process_name}'
        else:
            process_label = process_name

        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col)
        cell.value = process_label
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=cell_color,
                                end_color=cell_color, fill_type='solid')

        row2_color = lighten_color(cell_color)
        for i, step_item in enumerate(steps):
            col_index = start_col + i
            if isinstance(step_item, tuple):
                step_label, test_val = step_item
                cell = ws.cell(row=2, column=col_index)
                cell.value = step_label
                cell.fill = PatternFill(start_color=row2_color,
                                        end_color=row2_color, fill_type='solid')
                if is_testing:
                    ws.cell(row=3, column=col_index, value=test_val)
        start_col = end_col + 1
        incremental_number += 1

    # Example: Apply a custom formula for the "Nomad ID" column (example only)
    for row in range(3, 4):
        nomad_id_formula = f'=CONCATENATE("HZB_",B{row},"_",C{row},"_",D{row},"_C-",E{row})'
        ws[f'F{row}'].value = nomad_id_formula

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value and isinstance(cell.value, str):
                max_length = max(max_length, len(cell.value))
        ws.column_dimensions[column_letter].width = max_length + 2
